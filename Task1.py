import openpyxl
import xlrd 
from openpyxl import load_workbook
wb_obj = load_workbook("Task1_file.xlsx")
sheet_obj = wb_obj.active
 
data = []
dict_1 = {"AA" : "ECE",

"AB" : "Manufacturing",

"A1" : "Chemical",

"A2" : "Civil",

"A3" : "EEE",

"A4" : "Mech",

"A5" : "Pharma",

"A7" : "CSE",

"A8" : "ENI",

"B1" : "MSc BIO",

"B2" : "MSc Chem",

"B3" : "MSc Eco",

"B4" : "MSc Mathematics",

"B5" : "MSc Physics"}

for row in range (2, sheet_obj.max_row+1):
    name = sheet_obj.cell(row , 1).value
    id = sheet_obj.cell(row , 2).value
    if not id:
        break
    Branch_1 = dict_1.get(id[4:6])
    
    Branch_2 = dict_1.get(id[6:8], "None")
    email  = "f"+id[0:4]+id[8:12]+"@pilani.bits-pilani.ac.in"
    
    dict_2 = {"Name": name,
              "BITS ID": id,
              "Branch 1": Branch_1,
              "Branch 2": Branch_2,
              "Email": email}
    data.append(dict_2) 
    
print(data)

"""
.navbar-default, .btn, .form-control, .panel {
	border-radius: 0;
}
.navbar-default {
    background-color: rgba(245, 245, 220, .7);
    border-color: rgba(245, 245, 220, .6);
    box-shadow: 0px 3px 3px 0px #ddd;
}
.navbar-default .navbar-nav>.active>a, 
.navbar-default .navbar-nav>.active>a:focus, 
.navbar-default .navbar-nav>.active>a:hover {
    background-color: rgba(245, 245, 220, .9);
    border-bottom: 4px solid #ddd;
}
.panel-default > .panel-heading {
    color: #333;
    background-color: rgba(245, 245, 220, .7);
    border-color: rgba(245, 245, 220, .6);
    border-top-left-radius: 0;
	border-top-right-radius: 0; 
}
.panel-footer {
	background-color: rgba(245, 245, 220, .5);
    border-color: rgba(245, 245, 220, .4);
}
.panel {
	box-shadow: 0px 3px 3px 0px #ddd;
}
a.navbar-brand, .panel-title {
	font-weight: bolder;
}
"""    